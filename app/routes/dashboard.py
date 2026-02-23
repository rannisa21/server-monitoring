from flask import Blueprint, render_template, request, redirect, url_for, current_app, send_file, flash, session, jsonify
from flask_login import login_required, current_user
from app.models.server import Server, Component
from app.models.metric import Metric, wib_now
from app import db
from sqlalchemy import desc, asc, extract
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from collections import defaultdict
import logging

logger = logging.getLogger(__name__)

dashboard_bp = Blueprint('dashboard', __name__)

@dashboard_bp.route('/')
@login_required
def dashboard():
    try:
        # Get filter parameters
        server_filter = request.args.getlist('server_id', type=int)  # Multiple server filter
        category_filter = request.args.get('category', '')
        status_filter = request.args.get('status', '')
        search_query = request.args.get('search', '').strip()
        
        # Get view type (table or card), default to table
        view_type = request.args.get('view', session.get('dashboard_view', 'table'))
        session['dashboard_view'] = view_type
        
        # Get sort parameters
        sort_by = request.args.get('sort', 'server')  # Default sort by server
        sort_order = request.args.get('order', 'asc')  # Default ascending
        
        # Get all servers for filter dropdown
        all_servers = Server.query.order_by(Server.name).all()
        
        # Build servers query
        if server_filter:
            servers = Server.query.filter(Server.id.in_(server_filter)).all()
        else:
            servers = all_servers  # Show all by default
        
        # Build data for dashboard with latest metrics
        dashboard_data = []
        for server in servers:
            for component in server.components:
                metric = Metric.query.filter_by(
                    server_id=server.id,
                    component_id=component.id
                ).order_by(desc(Metric.timestamp)).first()
                
                # Apply category filter
                if category_filter and component.category != category_filter:
                    continue
                
                # Apply status filter
                if status_filter:
                    if status_filter == 'no_data' and metric:
                        continue
                    elif status_filter != 'no_data' and (not metric or metric.status != status_filter):
                        continue
                
                # Apply search filter
                if search_query:
                    search_lower = search_query.lower()
                    if not (search_lower in server.name.lower() or 
                            search_lower in component.name.lower() or
                            search_lower in server.ip.lower() or
                            search_lower in component.oid.lower()):
                        continue
                
                dashboard_data.append({
                    'server': server,
                    'component': component,
                    'metric': metric
                })
        
        # Sort data
        def get_sort_key(item):
            if sort_by == 'server':
                return item['server'].name.lower()
            elif sort_by == 'component':
                return item['component'].name.lower()
            elif sort_by == 'category':
                return item['component'].category.lower()
            elif sort_by == 'status':
                return item['metric'].status if item['metric'] else 'zzz'
            elif sort_by == 'timestamp':
                return item['metric'].timestamp if item['metric'] else datetime.min
            return item['server'].name.lower()
        
        dashboard_data.sort(key=get_sort_key, reverse=(sort_order == 'desc'))
        
        # Get unique categories for filter
        categories = db.session.query(Component.category).distinct().all()
        categories = [c[0] for c in categories]
        
        total_items = len(dashboard_data)
        
        # For card view, group by server
        card_data = {}
        if view_type == 'card':
            for item in dashboard_data:
                server_id = item['server'].id
                if server_id not in card_data:
                    card_data[server_id] = {
                        'server': item['server'],
                        'components': []
                    }
                card_data[server_id]['components'].append(item)
        
        logger.debug(f'Dashboard loaded: {total_items} items, view: {view_type}')
        
        return render_template(
            'dashboard.html',
            dashboard_data=dashboard_data,
            card_data=card_data,
            view_type=view_type,
            servers=all_servers,
            server_filter=server_filter,
            category_filter=category_filter,
            status_filter=status_filter,
            search_query=search_query,
            sort_by=sort_by,
            sort_order=sort_order,
            categories=categories,
            total_items=total_items
        )
        
    except Exception as e:
        logger.error(f'Dashboard error: {e}', exc_info=True)
        return render_template(
            'dashboard.html',
            dashboard_data=[],
            card_data={},
            view_type='table',
            servers=[],
            server_filter=[],
            category_filter='',
            status_filter='',
            search_query='',
            sort_by='server',
            sort_order='asc',
            categories=[],
            total_items=0,
            error='Failed to load dashboard data'
        )


@dashboard_bp.route('/api/data')
@login_required
def api_dashboard_data():
    """API endpoint untuk mendapatkan data dashboard terbaru (JSON)."""
    try:
        # Get filter parameters
        server_filter = request.args.getlist('server_id', type=int)
        category_filter = request.args.get('category', '')
        status_filter = request.args.get('status', '')
        search_query = request.args.get('search', '').strip()
        sort_by = request.args.get('sort', 'server')
        sort_order = request.args.get('order', 'asc')
        
        # Get all servers
        all_servers = Server.query.order_by(Server.name).all()
        
        # Build servers query
        if server_filter:
            servers = Server.query.filter(Server.id.in_(server_filter)).all()
        else:
            servers = all_servers
        
        # Build data for dashboard with latest metrics
        dashboard_data = []
        for server in servers:
            for component in server.components:
                metric = Metric.query.filter_by(
                    server_id=server.id,
                    component_id=component.id
                ).order_by(desc(Metric.timestamp)).first()
                
                # Apply category filter
                if category_filter and component.category != category_filter:
                    continue
                
                # Apply status filter
                if status_filter:
                    if status_filter == 'no_data' and metric:
                        continue
                    elif status_filter != 'no_data' and (not metric or metric.status != status_filter):
                        continue
                
                # Apply search filter
                if search_query:
                    search_lower = search_query.lower()
                    if not (search_lower in server.name.lower() or 
                            search_lower in component.name.lower() or
                            search_lower in server.ip.lower() or
                            search_lower in component.oid.lower()):
                        continue
                
                dashboard_data.append({
                    'server_id': server.id,
                    'server_name': server.name,
                    'server_ip': server.ip,
                    'server_brand': server.brand,
                    'component_id': component.id,
                    'component_name': component.name,
                    'component_oid': component.oid,
                    'category': component.category,
                    'metric_value': metric.value if metric else None,
                    'metric_status': metric.status if metric else None,
                    'metric_timestamp': metric.timestamp.strftime('%Y-%m-%d %H:%M:%S') if metric else None
                })
        
        # Sort data
        def get_sort_key(item):
            if sort_by == 'server':
                return item['server_name'].lower()
            elif sort_by == 'component':
                return item['component_name'].lower()
            elif sort_by == 'category':
                return item['category'].lower()
            elif sort_by == 'status':
                return item['metric_status'] if item['metric_status'] else 'zzz'
            elif sort_by == 'timestamp':
                return item['metric_timestamp'] if item['metric_timestamp'] else ''
            return item['server_name'].lower()
        
        dashboard_data.sort(key=get_sort_key, reverse=(sort_order == 'desc'))
        
        # Get last update time from newest metric
        latest_metric = Metric.query.order_by(desc(Metric.timestamp)).first()
        last_update = latest_metric.timestamp.strftime('%Y-%m-%d %H:%M:%S') if latest_metric else None
        
        return jsonify({
            'success': True,
            'data': dashboard_data,
            'total_items': len(dashboard_data),
            'last_update': last_update
        })
        
    except Exception as e:
        logger.error(f'API Dashboard error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e),
            'data': [],
            'total_items': 0
        }), 500


def calculate_component_uptime(metrics):
    """
    Calculate uptime percentage based on status periods.
    A period is a contiguous time span where the status is consistently OK or non-OK.
    Similar to period analysis logic in download report.
    Uptime % = OK_periods / total_periods * 100
    """
    if not metrics:
        return None, None

    ok_periods = 0
    non_ok_periods = 0
    prev_is_ok = None
    current_status = None

    for m in metrics:
        is_ok = (m.status == 'OK')
        if prev_is_ok is None or is_ok != prev_is_ok:
            if is_ok:
                ok_periods += 1
            else:
                non_ok_periods += 1
        prev_is_ok = is_ok
        current_status = m.status

    total_periods = ok_periods + non_ok_periods
    if total_periods == 0:
        return 0.0, current_status

    uptime_pct = round((ok_periods / total_periods) * 100, 2)
    return uptime_pct, current_status


@dashboard_bp.route('/api/uptime')
@login_required
def api_uptime_data():
    """API endpoint to get uptime percentages per component per server."""
    try:
        mode = request.args.get('mode', 'quick')
        now = wib_now()

        if mode == 'quick':
            amount = request.args.get('amount', 1, type=int)
            unit = request.args.get('unit', 'hour')
            if amount < 1:
                amount = 1
            delta_map = {
                'minute': timedelta(minutes=amount),
                'hour': timedelta(hours=amount),
                'day': timedelta(days=amount),
                'week': timedelta(weeks=amount),
                'month': timedelta(days=amount * 30),
                'year': timedelta(days=amount * 365),
            }
            delta = delta_map.get(unit, timedelta(hours=1))
            from_time = now - delta
            to_time = now
        else:
            from_str = request.args.get('from_datetime', '')
            to_str = request.args.get('to_datetime', '')
            try:
                from_time = datetime.strptime(from_str, '%Y-%m-%dT%H:%M')
                to_time = datetime.strptime(to_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                return jsonify({'success': False, 'error': 'Format datetime tidak valid'}), 400
            if from_time >= to_time:
                return jsonify({'success': False, 'error': 'Waktu mulai harus sebelum waktu selesai'}), 400

        server_filter = request.args.getlist('server_id', type=int)
        category_filter = request.args.get('category', '').strip()
        status_filter = request.args.get('status', '').strip()
        search_query = request.args.get('search', '').strip()
        sort_by = request.args.get('sort', 'server')
        sort_order = request.args.get('order', 'asc')

        if server_filter:
            servers = Server.query.filter(Server.id.in_(server_filter)).all()
        else:
            servers = Server.query.order_by(Server.name).all()

        uptime_data = {}
        flat_items = []  # collect flat list for sorting
        for server in servers:
            for component in server.components:
                # Apply category filter
                if category_filter and component.category != category_filter:
                    continue

                # Apply search filter
                if search_query:
                    sl = search_query.lower()
                    if not (sl in server.name.lower() or
                            sl in component.name.lower() or
                            sl in server.ip.lower() or
                            sl in component.oid.lower()):
                        continue

                metrics = Metric.query.filter(
                    Metric.server_id == server.id,
                    Metric.component_id == component.id,
                    Metric.timestamp >= from_time,
                    Metric.timestamp <= to_time
                ).order_by(Metric.timestamp.asc()).all()

                uptime_pct, current_status = calculate_component_uptime(metrics)

                # Apply status filter
                if status_filter:
                    if status_filter == 'no_data' and current_status is not None:
                        continue
                    elif status_filter != 'no_data' and current_status != status_filter:
                        continue

                flat_items.append({
                    'server_id': str(server.id),
                    'server_name': server.name,
                    'server_ip': server.ip,
                    'component_id': component.id,
                    'component_name': component.name,
                    'category': component.category,
                    'uptime_pct': uptime_pct,
                    'current_status': current_status,
                    'total_metrics': len(metrics)
                })

        # Sort flat items
        def sort_key(item):
            if sort_by == 'server':
                return item['server_name'].lower()
            elif sort_by == 'component':
                return item['component_name'].lower()
            elif sort_by == 'category':
                return item['category'].lower()
            elif sort_by == 'status':
                return item['current_status'] or 'zzz'
            elif sort_by == 'uptime':
                return item['uptime_pct'] if item['uptime_pct'] is not None else -1
            return item['server_name'].lower()

        flat_items.sort(key=sort_key, reverse=(sort_order == 'desc'))

        # Group back into server-keyed structure (preserve sort order)
        from collections import OrderedDict
        uptime_data = OrderedDict()
        for item in flat_items:
            sid = item['server_id']
            if sid not in uptime_data:
                uptime_data[sid] = {
                    'server_name': item['server_name'],
                    'server_ip': item['server_ip'],
                    'components': []
                }
            uptime_data[sid]['components'].append({
                'component_id': item['component_id'],
                'component_name': item['component_name'],
                'category': item['category'],
                'uptime_pct': item['uptime_pct'],
                'current_status': item['current_status'],
                'total_metrics': item['total_metrics']
            })

        return jsonify({
            'success': True,
            'data': uptime_data,
            'total_items': len(flat_items),
            'from_time': from_time.strftime('%Y-%m-%d %H:%M:%S'),
            'to_time': to_time.strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        logger.error(f'Uptime API error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@dashboard_bp.route('/download-report', methods=['POST'])
@login_required
def download_report():
    """Download Excel report with 3 sheets. Supports month, quick period, and custom range."""
    try:
        report_mode = request.form.get('report_mode', 'month')

        if report_mode == 'quick':
            amount = request.form.get('report_amount', type=int)
            unit = request.form.get('report_unit', 'hour')
            if not amount or amount < 1:
                flash('Masukkan jumlah periode yang valid.', 'danger')
                return redirect(url_for('dashboard.dashboard'))
            now = wib_now()
            delta_map = {
                'minute': timedelta(minutes=amount),
                'hour': timedelta(hours=amount),
                'day': timedelta(days=amount),
                'week': timedelta(weeks=amount),
                'month': timedelta(days=amount * 30),
                'year': timedelta(days=amount * 365),
            }
            unit_labels = {'minute': 'Menit', 'hour': 'Jam', 'day': 'Hari', 'week': 'Minggu', 'month': 'Bulan', 'year': 'Tahun'}
            delta = delta_map.get(unit, timedelta(hours=1))
            from_time = now - delta
            to_time = now
            metrics = Metric.query.filter(
                Metric.timestamp >= from_time,
                Metric.timestamp <= to_time
            ).order_by(Metric.timestamp.asc()).all()
            period_label = f"Last {amount} {unit_labels.get(unit, unit)}"
            filename = f"laporan_monitoring_last_{amount}_{unit}.xlsx"

        elif report_mode == 'range':
            from_str = request.form.get('report_from', '')
            to_str = request.form.get('report_to', '')
            try:
                from_time = datetime.strptime(from_str, '%Y-%m-%dT%H:%M')
                to_time = datetime.strptime(to_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Format datetime tidak valid.', 'danger')
                return redirect(url_for('dashboard.dashboard'))
            if from_time >= to_time:
                flash('Waktu mulai harus sebelum waktu selesai.', 'danger')
                return redirect(url_for('dashboard.dashboard'))
            metrics = Metric.query.filter(
                Metric.timestamp >= from_time,
                Metric.timestamp <= to_time
            ).order_by(Metric.timestamp.asc()).all()
            period_label = f"{from_time.strftime('%Y-%m-%d %H:%M')} s/d {to_time.strftime('%Y-%m-%d %H:%M')}"
            filename = f"laporan_monitoring_{from_time.strftime('%Y%m%d_%H%M')}_{to_time.strftime('%Y%m%d_%H%M')}.xlsx"

        else:  # month mode (default)
            month = request.form.get('month', type=int)
            year = request.form.get('year', type=int)
            if not month or not year:
                flash('Pilih bulan dan tahun terlebih dahulu.', 'danger')
                return redirect(url_for('dashboard.dashboard'))
            if month < 1 or month > 12:
                flash('Bulan tidak valid.', 'danger')
                return redirect(url_for('dashboard.dashboard'))
            metrics = Metric.query.filter(
                extract('month', Metric.timestamp) == month,
                extract('year', Metric.timestamp) == year
            ).order_by(Metric.timestamp.asc()).all()
            period_label = f"{month:02d}/{year}"
            filename = f"laporan_monitoring_{year}_{month:02d}.xlsx"

        if not metrics:
            flash(f'Tidak ada data untuk periode {period_label}.', 'warning')
            return redirect(url_for('dashboard.dashboard'))

        # Get server SNMP versions
        server_snmp_versions = {}
        servers = Server.query.all()
        for s in servers:
            server_snmp_versions[s.name] = s.snmp_version
        
        # ============ SHEET 1: Raw Data ============
        raw_data = []
        for i, m in enumerate(metrics):
            snmp_version = server_snmp_versions.get(m.server_name, 'N/A')
            raw_data.append({
                'No': i + 1,
                'Server Name': m.server_name,
                'IP Address': m.server_ip,
                'Brand': m.brand,
                'SNMP Version': snmp_version,
                'Category': m.category,
                'Component Name': m.component_name,
                'OID': m.oid,
                'Value': m.value,
                'Timestamp': m.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'Status': m.status
            })
        
        df_raw = pd.DataFrame(raw_data)
        
        # ============ SHEET 2: Period Analysis ============
        # Group metrics by server and component
        grouped = defaultdict(list)
        for m in raw_data:
            key = (m['Server Name'], m['IP Address'], m['Component Name'], m['Category'])
            grouped[key].append(m)
        
        analysis_results = []
        for (server_name, ip_address, component_name, category), records in grouped.items():
            # Sort by timestamp ascending for proper period calculation
            sorted_records = sorted(records, key=lambda x: x['Timestamp'])
            
            # Track periods - count when status CHANGES to that status
            periods = {
                'OK': [],
                'Warning': [],
                'Critical': [],
                'Failed': []
            }
            
            prev_status = None
            for record in sorted_records:
                current_status = record['Status']
                timestamp = record['Timestamp']
                
                # If status changed, record the new period
                if current_status != prev_status:
                    if current_status in periods:
                        periods[current_status].append(timestamp)
                
                prev_status = current_status
            
            # Build analysis row
            analysis_row = {
                'Server Name': server_name,
                'IP Address': ip_address,
                'Component Name': component_name,
                'Category': category,
                'Total OK Periods': len(periods['OK']),
                'Total Warning Periods': len(periods['Warning']),
                'Total Critical Periods': len(periods['Critical']),
                'Total Failed Periods': len(periods['Failed']),
                'Critical Period Timestamps': ', '.join(periods['Critical']) if periods['Critical'] else '-',
                'Warning Period Timestamps': ', '.join(periods['Warning']) if periods['Warning'] else '-',
                'Failed Period Timestamps': ', '.join(periods['Failed']) if periods['Failed'] else '-',
                'Total Records': len(sorted_records)
            }
            analysis_results.append(analysis_row)
        
        df_analysis = pd.DataFrame(analysis_results)
        
        # ============ SHEET 3: Server Summary for Visualization ============
        server_summary = defaultdict(lambda: {'OK': 0, 'Warning': 0, 'Critical': 0, 'Failed': 0})
        for row in analysis_results:
            server_name = row['Server Name']
            server_summary[server_name]['OK'] += row['Total OK Periods']
            server_summary[server_name]['Warning'] += row['Total Warning Periods']
            server_summary[server_name]['Critical'] += row['Total Critical Periods']
            server_summary[server_name]['Failed'] += row['Total Failed Periods']
        
        # ============ CREATE EXCEL FILE ============
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # --- Sheet 1: Raw Data ---
            df_raw.to_excel(writer, index=False, sheet_name='Raw Data')
            ws_raw = writer.sheets['Raw Data']
            
            # Auto-adjust column widths
            for idx, col in enumerate(df_raw.columns):
                max_length = max(df_raw[col].astype(str).map(len).max(), len(str(col))) + 2
                ws_raw.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
            
            # Style header
            for col in range(1, len(df_raw.columns) + 1):
                cell = ws_raw.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # --- Sheet 2: Period Analysis ---
            df_analysis.to_excel(writer, index=False, sheet_name='Period Analysis')
            ws_analysis = writer.sheets['Period Analysis']
            
            # Auto-adjust column widths
            for idx, col in enumerate(df_analysis.columns):
                max_length = max(df_analysis[col].astype(str).map(len).max(), len(str(col))) + 2
                ws_analysis.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
            
            # Style header
            for col in range(1, len(df_analysis.columns) + 1):
                cell = ws_analysis.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Conditional formatting for critical/warning counts
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            if len(df_analysis) > 0:
                critical_col = get_column_letter(df_analysis.columns.get_loc('Total Critical Periods') + 1)
                warning_col = get_column_letter(df_analysis.columns.get_loc('Total Warning Periods') + 1)
                data_rows = len(df_analysis) + 1
                
                ws_analysis.conditional_formatting.add(
                    f'{critical_col}2:{critical_col}{data_rows}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill)
                )
                ws_analysis.conditional_formatting.add(
                    f'{warning_col}2:{warning_col}{data_rows}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=orange_fill)
                )
            
            # --- Sheet 3: Visualization (Uptime % per component per server) ---
            workbook = writer.book
            ws_viz = workbook.create_sheet(title='Visualization')
            
            # Title
            ws_viz['A1'] = f'Component Uptime (OK %) per Server - {period_label}'
            ws_viz['A1'].font = Font(bold=True, size=14)
            ws_viz.merge_cells('A1:D1')
            
            # Calculate OK% per component per server using period logic
            server_components_uptime = defaultdict(list)  # server_name -> [(comp_name, ok_pct)]
            for (server_name, ip_address, component_name, category), records in grouped.items():
                sorted_recs = sorted(records, key=lambda x: x['Timestamp'])
                ok_periods = 0
                non_ok_periods = 0
                prev_is_ok = None
                for rec in sorted_recs:
                    is_ok = (rec['Status'] == 'OK')
                    if prev_is_ok is None or is_ok != prev_is_ok:
                        if is_ok:
                            ok_periods += 1
                        else:
                            non_ok_periods += 1
                    prev_is_ok = is_ok
                total_p = ok_periods + non_ok_periods
                ok_pct = round((ok_periods / total_p) * 100, 2) if total_p > 0 else 0.0
                server_components_uptime[server_name].append((component_name, ok_pct))
            
            # Styling helpers
            green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            red_fill_viz = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            ws_viz.column_dimensions['A'].width = 35
            ws_viz.column_dimensions['B'].width = 15
            
            current_row = 3
            chart_col_offset = 0  # For placing charts side by side
            chart_count = 0
            
            for server_name, comp_list in server_components_uptime.items():
                # Sort components alphabetically
                comp_list.sort(key=lambda x: x[0])
                
                # Write server section header
                ws_viz.cell(row=current_row, column=1, value=server_name)
                ws_viz.cell(row=current_row, column=1).font = Font(bold=True, size=11)
                current_row += 1
                
                # Write column headers
                hdr_comp = ws_viz.cell(row=current_row, column=1, value='Component')
                hdr_pct = ws_viz.cell(row=current_row, column=2, value='OK %')
                for hdr_cell in [hdr_comp, hdr_pct]:
                    hdr_cell.fill = header_fill
                    hdr_cell.font = header_font
                    hdr_cell.alignment = Alignment(horizontal='center')
                
                data_start_row = current_row  # header row (for chart titles_from_data)
                current_row += 1
                
                for comp_name, ok_pct in comp_list:
                    ws_viz.cell(row=current_row, column=1, value=comp_name)
                    pct_cell = ws_viz.cell(row=current_row, column=2, value=ok_pct)
                    pct_cell.number_format = '0.00'
                    pct_cell.alignment = Alignment(horizontal='center')
                    # Color the cell based on OK%
                    if ok_pct >= 90:
                        pct_cell.fill = green_fill
                    elif ok_pct >= 50:
                        pct_cell.fill = yellow_fill
                    else:
                        pct_cell.fill = red_fill_viz
                    current_row += 1
                
                data_end_row = current_row - 1
                
                # Create horizontal bar chart for this server
                if len(comp_list) > 0:
                    chart = BarChart()
                    chart.type = 'bar'  # horizontal bars
                    chart.grouping = 'clustered'
                    chart.title = f'{server_name} - Component Uptime (OK %)'
                    chart.x_axis.title = 'OK %'
                    chart.y_axis.title = None
                    chart.x_axis.scaling.min = 0
                    chart.x_axis.scaling.max = 100
                    chart.x_axis.majorUnit = 10
                    chart.legend = None
                    
                    data_ref = Reference(ws_viz, min_col=2, min_row=data_start_row, max_row=data_end_row)
                    cats_ref = Reference(ws_viz, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    
                    chart.width = 20
                    chart.height = max(6, len(comp_list) * 1.2 + 3)
                    
                    # Green bar color
                    if len(chart.series) >= 1:
                        chart.series[0].graphicalProperties.solidFill = '28A745'
                    
                    # Place charts: 2 columns of charts
                    chart_col = 'D' if chart_count % 2 == 0 else 'P'
                    chart_anchor_row = 3 + (chart_count // 2) * 20
                    ws_viz.add_chart(chart, f'{chart_col}{chart_anchor_row}')
                    chart_count += 1
                
                current_row += 1  # blank row between servers
            
            # Report info at the bottom
            info_row = current_row + 2
            ws_viz.cell(row=info_row, column=1, value=f"Report Period: {period_label}")
            ws_viz.cell(row=info_row, column=1).font = Font(italic=True)
            ws_viz.cell(row=info_row + 1, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws_viz.cell(row=info_row + 1, column=1).font = Font(italic=True)
            ws_viz.cell(row=info_row + 2, column=1, value=f"Generated by: {current_user.username}")
            ws_viz.cell(row=info_row + 2, column=1).font = Font(italic=True)
        
        output.seek(0)
        
        logger.info(f'Report downloaded from dashboard for {period_label} by {current_user.username}: {len(metrics)} records, 3 sheets')
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f'Error downloading report: {e}', exc_info=True)
        flash('Terjadi kesalahan saat mengunduh report.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
