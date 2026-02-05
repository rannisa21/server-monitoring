from flask import Blueprint, render_template, request, send_file, flash, current_app
from flask_login import login_required, current_user
from app.models.metric import Metric
from app.models.server import Server, Component
from io import BytesIO
import pandas as pd
from datetime import datetime
from app import db
from app.validators import admin_required, validate_month_year, ValidationError
from sqlalchemy import extract
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from collections import defaultdict
import logging

logger = logging.getLogger(__name__)

report_bp = Blueprint('report', __name__)


def calculate_status_periods(metrics_data):
    """
    Calculate status periods per server and component.
    A period is counted when status first changes (e.g., from OK to Critical/Warning),
    and not counted again while status remains the same until it changes.
    
    Returns a list of dictionaries with period analysis data.
    """
    # Group metrics by server and component
    grouped = defaultdict(list)
    for m in metrics_data:
        key = (m['Server Name'], m['IP Address'], m['Component Name'], m['Category'])
        grouped[key].append(m)
    
    analysis_results = []
    
    for (server_name, ip_address, component_name, category), records in grouped.items():
        # Sort by timestamp ascending for proper period calculation
        sorted_records = sorted(records, key=lambda x: x['Timestamp'])
        
        # Track periods
        periods = {
            'OK': [],
            'Warning': [],
            'Critical': [],
            'Failed': []
        }
        
        prev_status = None
        current_period_start = None
        
        for record in sorted_records:
            current_status = record['Status']
            timestamp = record['Timestamp']
            
            # If status changed, record the new period
            if current_status != prev_status:
                if current_status in periods:
                    periods[current_status].append(timestamp)
                current_period_start = timestamp
            
            prev_status = current_status
        
        # Build analysis row
        analysis_row = {
            'Server Name': server_name,
            'IP Address': ip_address,
            'Component Name': component_name,
            'Category': category,
            'Total OK Periods': len(periods['OK']),
            'Total Warning Periods': len(periods['Warning']),
            'Total Critical Periods': len(periods['Critical']),
            'Total Failed Periods': len(periods['Failed']),
            'Critical Period Timestamps': ', '.join([str(ts) for ts in periods['Critical']]) if periods['Critical'] else '-',
            'Warning Period Timestamps': ', '.join([str(ts) for ts in periods['Warning']]) if periods['Warning'] else '-',
            'Failed Period Timestamps': ', '.join([str(ts) for ts in periods['Failed']]) if periods['Failed'] else '-',
            'Total Records': len(sorted_records)
        }
        
        analysis_results.append(analysis_row)
    
    return analysis_results


def calculate_server_summary(analysis_data):
    """
    Calculate summary per server for visualization.
    """
    server_summary = defaultdict(lambda: {
        'OK': 0, 'Warning': 0, 'Critical': 0, 'Failed': 0
    })
    
    for row in analysis_data:
        server_name = row['Server Name']
        server_summary[server_name]['OK'] += row['Total OK Periods']
        server_summary[server_name]['Warning'] += row['Total Warning Periods']
        server_summary[server_name]['Critical'] += row['Total Critical Periods']
        server_summary[server_name]['Failed'] += row['Total Failed Periods']
    
    return server_summary


def auto_adjust_column_width(worksheet, df):
    """Auto-adjust column widths based on content."""
    for idx, col in enumerate(df.columns):
        max_length = max(
            df[col].astype(str).map(len).max(),
            len(str(col))
        ) + 2
        column_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[column_letter].width = min(max_length, 50)


def style_header(worksheet, num_cols):
    """Apply styling to header row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def create_visualization_sheet(workbook, server_summary, month, year):
    """Create visualization sheet with charts."""
    # Create data for charts
    ws_viz = workbook.create_sheet(title='Visualization')
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write summary data table for chart
    ws_viz['A1'] = 'Server Status Period Summary'
    ws_viz['A1'].font = Font(bold=True, size=14)
    ws_viz.merge_cells('A1:E1')
    
    # Headers
    headers = ['Server Name', 'OK Periods', 'Warning Periods', 'Critical Periods', 'Failed Periods']
    for col, header in enumerate(headers, 1):
        cell = ws_viz.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 4
    for server_name, counts in server_summary.items():
        ws_viz.cell(row=row, column=1, value=server_name)
        ws_viz.cell(row=row, column=2, value=counts['OK'])
        ws_viz.cell(row=row, column=3, value=counts['Warning'])
        ws_viz.cell(row=row, column=4, value=counts['Critical'])
        ws_viz.cell(row=row, column=5, value=counts['Failed'])
        row += 1
    
    # Adjust column widths
    ws_viz.column_dimensions['A'].width = 30
    ws_viz.column_dimensions['B'].width = 15
    ws_viz.column_dimensions['C'].width = 18
    ws_viz.column_dimensions['D'].width = 18
    ws_viz.column_dimensions['E'].width = 15
    
    # Create Bar Chart for Critical/Warning/Failed periods per server
    if len(server_summary) > 0:
        chart1 = BarChart()
        chart1.type = "col"
        chart1.grouping = "clustered"
        chart1.title = f"Status Periods per Server ({month:02d}/{year})"
        chart1.y_axis.title = "Number of Periods"
        chart1.x_axis.title = "Server"
        
        data_end_row = 3 + len(server_summary)
        
        # Data references (Warning, Critical, Failed)
        data = Reference(ws_viz, min_col=3, min_row=3, max_col=5, max_row=data_end_row)
        cats = Reference(ws_viz, min_col=1, min_row=4, max_row=data_end_row)
        
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        chart1.width = 18
        chart1.height = 10
        
        # Color the series
        if len(chart1.series) >= 1:
            chart1.series[0].graphicalProperties.solidFill = "FFC000"  # Warning - Orange
        if len(chart1.series) >= 2:
            chart1.series[1].graphicalProperties.solidFill = "FF0000"  # Critical - Red
        if len(chart1.series) >= 3:
            chart1.series[2].graphicalProperties.solidFill = "7030A0"  # Failed - Purple
        
        ws_viz.add_chart(chart1, "G3")
        
        # Create second chart - Total Issues per Server (Stacked)
        chart2 = BarChart()
        chart2.type = "col"
        chart2.grouping = "stacked"
        chart2.title = f"Total Issue Periods per Server ({month:02d}/{year})"
        chart2.y_axis.title = "Total Periods"
        chart2.x_axis.title = "Server"
        
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.width = 18
        chart2.height = 10
        
        # Color the series
        if len(chart2.series) >= 1:
            chart2.series[0].graphicalProperties.solidFill = "FFC000"  # Warning - Orange
        if len(chart2.series) >= 2:
            chart2.series[1].graphicalProperties.solidFill = "FF0000"  # Critical - Red
        if len(chart2.series) >= 3:
            chart2.series[2].graphicalProperties.solidFill = "7030A0"  # Failed - Purple
        
        ws_viz.add_chart(chart2, "G20")
    
    # Add report info
    info_row = max(row + 2, 38)
    ws_viz.cell(row=info_row, column=1, value=f"Report Period: {month:02d}/{year}")
    ws_viz.cell(row=info_row, column=1).font = Font(italic=True)
    ws_viz.cell(row=info_row + 1, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws_viz.cell(row=info_row + 1, column=1).font = Font(italic=True)
    
    return ws_viz


@report_bp.route('/admin/report', methods=['GET', 'POST'])
@login_required
@admin_required
def report():
    if request.method == 'POST':
        try:
            # Validate inputs
            month, year = validate_month_year(
                request.form.get('month'),
                request.form.get('year')
            )
            
            # Query metrics for the specified month/year with server info
            metrics = Metric.query.filter(
                extract('month', Metric.timestamp) == month,
                extract('year', Metric.timestamp) == year
            ).order_by(Metric.timestamp.asc()).all()
            
            if not metrics:
                flash(f'No data found for {month:02d}/{year}.', 'warning')
                return render_template('report.html')
            
            # Get server info for SNMP version
            server_snmp_versions = {}
            servers = Server.query.all()
            for s in servers:
                server_snmp_versions[s.name] = s.snmp_version
            
            # Build raw data for Sheet 1
            raw_data = []
            for i, m in enumerate(metrics):
                snmp_version = server_snmp_versions.get(m.server_name, 'N/A')
                raw_data.append({
                    'No': i + 1,
                    'Server Name': m.server_name,
                    'IP Address': m.server_ip,
                    'Brand': m.brand,
                    'SNMP Version': snmp_version,
                    'Category': m.category,
                    'Component Name': m.component_name,
                    'OID': m.oid,
                    'Value': m.value,
                    'Timestamp': m.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': m.status
                })
            
            df_raw = pd.DataFrame(raw_data)
            
            # Calculate period analysis for Sheet 2
            analysis_data = calculate_status_periods(raw_data)
            df_analysis = pd.DataFrame(analysis_data)
            
            # Calculate server summary for visualization
            server_summary = calculate_server_summary(analysis_data)
            
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: Raw Data
                df_raw.to_excel(writer, index=False, sheet_name='Raw Data')
                ws_raw = writer.sheets['Raw Data']
                auto_adjust_column_width(ws_raw, df_raw)
                style_header(ws_raw, len(df_raw.columns))
                
                # Sheet 2: Period Analysis
                df_analysis.to_excel(writer, index=False, sheet_name='Period Analysis')
                ws_analysis = writer.sheets['Period Analysis']
                auto_adjust_column_width(ws_analysis, df_analysis)
                style_header(ws_analysis, len(df_analysis.columns))
                
                # Apply conditional formatting for critical/warning counts
                from openpyxl.formatting.rule import CellIsRule
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                purple_fill = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")
                
                # Get column letters for conditional formatting
                critical_col = get_column_letter(df_analysis.columns.get_loc('Total Critical Periods') + 1)
                warning_col = get_column_letter(df_analysis.columns.get_loc('Total Warning Periods') + 1)
                failed_col = get_column_letter(df_analysis.columns.get_loc('Total Failed Periods') + 1)
                
                data_rows = len(df_analysis) + 1
                
                # Add conditional formatting rules
                ws_analysis.conditional_formatting.add(
                    f'{critical_col}2:{critical_col}{data_rows}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill)
                )
                ws_analysis.conditional_formatting.add(
                    f'{warning_col}2:{warning_col}{data_rows}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=orange_fill)
                )
                ws_analysis.conditional_formatting.add(
                    f'{failed_col}2:{failed_col}{data_rows}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=purple_fill)
                )
                
                # Sheet 3: Visualization
                workbook = writer.book
                create_visualization_sheet(workbook, server_summary, month, year)
            
            output.seek(0)
            filename = f"report_{year}_{month:02d}.xlsx"
            
            logger.info(f'Report generated for {month:02d}/{year} by {current_user.username}: {len(metrics)} records')
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except ValidationError as e:
            flash(e.message, 'danger')
            logger.warning(f'Validation error generating report: {e.message}')
        except Exception as e:
            logger.error(f'Error generating report: {e}', exc_info=True)
            flash('An error occurred while generating the report.', 'danger')
    
    return render_template('report.html')


@report_bp.route('/admin/report/preview', methods=['GET'])
@login_required
@admin_required
def report_preview():
    """Preview metrics with pagination."""
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = current_app.config.get('ITEMS_PER_PAGE', 20)
        
        if month and year:
            try:
                month, year = validate_month_year(month, year)
                
                pagination = Metric.query.filter(
                    extract('month', Metric.timestamp) == month,
                    extract('year', Metric.timestamp) == year
                ).order_by(Metric.timestamp.desc()).paginate(
                    page=page, per_page=per_page, error_out=False
                )
                
                return render_template(
                    'report.html',
                    metrics=pagination.items,
                    pagination=pagination,
                    month=month,
                    year=year
                )
            except ValidationError as e:
                flash(e.message, 'danger')
        
        return render_template('report.html')
        
    except Exception as e:
        logger.error(f'Error loading report preview: {e}', exc_info=True)
        flash('An error occurred while loading the report preview.', 'danger')
        return render_template('report.html')
